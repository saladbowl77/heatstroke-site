import openpyxl

wb = openpyxl.load_workbook('/Users/taiyo/Desktop/Project/tanpatu/heatstroke-site/python/heatstroke003_data_r4.xlsx')
sheet = wb['2022_05']

day = 31

dataArr = []

for num in range(47):
    dataArr.append([])

print(len(dataArr))

for rowNum in range(47 * day):
    pref = sheet.cell(row=rowNum + 2,column=2).value
    day = sheet.cell(row=rowNum + 2,column=1).value

    dataArr[int(pref - 1)].append([
        day.strftime('%Y/%m/%d'),
        sheet.cell(row=rowNum + 2,column=3).value,
        sheet.cell(row=rowNum + 2,column=4).value,
        sheet.cell(row=rowNum + 2,column=5).value,
        sheet.cell(row=rowNum + 2,column=6).value,
        sheet.cell(row=rowNum + 2,column=7).value,
        sheet.cell(row=rowNum + 2,column=8).value,
        sheet.cell(row=rowNum + 2,column=9).value,
        sheet.cell(row=rowNum + 2,column=10).value,
        sheet.cell(row=rowNum + 2,column=11).value,
        sheet.cell(row=rowNum + 2,column=12).value,
        sheet.cell(row=rowNum + 2,column=13).value,
        sheet.cell(row=rowNum + 2,column=14).value,
        sheet.cell(row=rowNum + 2,column=15).value,
        sheet.cell(row=rowNum + 2,column=16).value,
        sheet.cell(row=rowNum + 2,column=17).value,
        sheet.cell(row=rowNum + 2,column=18).value,
        sheet.cell(row=rowNum + 2,column=19).value,
        sheet.cell(row=rowNum + 2,column=20).value,
        sheet.cell(row=rowNum + 2,column=21).value,
        sheet.cell(row=rowNum + 2,column=22).value
    ])

import pandas as pd

prefNumber = {
    0 : "全国",
    1 : "北海道",
    2 : "青森県",
    3 : "岩手県",
    4 : "宮城県",
    5 : "秋田県",
    6 : "山形県",
    7 : "福島県",
    8 : "茨城県",
    9 : "栃木県",
    10 : "群馬県",
    11 : "埼玉県",
    12 : "千葉県",
    13 : "東京都",
    14 : "神奈川県",
    15 : "新潟県",
    16 : "富山県",
    17 : "石川県",
    18 : "福島県",
    19 : "山梨県",
    20 : "長野県",
    21 : "岐阜県",
    22 : "静岡県",
    23 : "愛知県",
    24 : "三重県",
    25 : "滋賀県",
    26 : "京都府",
    27 : "大阪府",
    28 : "兵庫県",
    29 : "奈良県",
    30 : "和歌山県",
    31 : "鳥取県",
    32 : "島根県",
    33 : "岡山県",
    34 : "広島県",
    35 : "山口県",
    36 : "徳島県",
    37 : "香川県",
    38 : "愛媛県",
    39 : "高知県",
    40 : "福岡県",
    41 : "佐賀県",
    42 : "長崎県",
    43 : "熊本県",
    44 : "大分県",
    45 : "宮崎県",
    46 : "鹿児島県",
    47 : "沖縄県",
}

prefCount=1
for dataPref in dataArr:
    df = pd.DataFrame(dataPref)
    df.to_csv(f'../src/lib/data/05/{prefCount}.csv', header=False, index=False)
    prefCount += 1